"""
Executor Module — Intent-based execution with Direct and UI modes.
Includes Action Guard Layer and Confidence Scoring.
"""

import os
import time
import subprocess
import platform
from typing import Any, Dict, Optional, Tuple

import openpyxl
from openpyxl import Workbook
import pyautogui

from core.logger import AgentLogger
from models.data_models import Plan, PlanStep, ExecutionMode, StructuredTask
from utils.excel_formatter import ExcelFormatter
from modules.screen_analyzer import ScreenAnalyzer
from modules.visual_executor import VisualExecutor


class DirectExecutor:
    """Execute plan intents directly using openpyxl."""

    def __init__(self, config: dict, logger: AgentLogger):
        self.config = config
        self.logger = logger
        self.output_dir = config.get("paths", {}).get("output_dir", "output")
        os.makedirs(self.output_dir, exist_ok=True)
        self.formatter = ExcelFormatter(config)
        self.workbook: Optional[Workbook] = None
        self.worksheet = None
        self.current_file: Optional[str] = None

    def execute_step(self, step: PlanStep, task: StructuredTask) -> Tuple[bool, str, float]:
        """Returns: (success, message, confidence)"""
        intent = step.intent
        params = step.params
        self.logger.info(f"[DIRECT] Executing intent: {intent} → {step.target}")

        try:
            if intent == "create_workbook":
                return self._create_workbook(params)
            elif intent == "write_headers":
                return self._write_headers(params, task)
            elif intent == "write_data":
                return self._write_data(params, task)
            elif intent == "format_cells":
                return self._format_cells(params, task)
            elif intent == "save_file":
                return self._save_file(params, task)
            else:
                return False, f"Unknown intent: {intent}", 0.0
        except Exception as e:
            error_msg = f"Direct execution error in '{intent}': {str(e)}"
            self.logger.error(error_msg)
            return False, error_msg, 0.0

    def _create_workbook(self, params: dict) -> Tuple[bool, str, float]:
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Data"
        filename = params.get("filename", "output.xlsx")
        self.current_file = os.path.join(self.output_dir, filename)
        return True, f"Created workbook: {self.current_file}", 1.0

    def _write_headers(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self.worksheet: return False, "No active worksheet", 0.0
        columns = params.get("columns", task.columns)
        for col_idx, header in enumerate(columns, 1):
            self.worksheet.cell(row=1, column=col_idx, value=str(header))
        return True, f"Added {len(columns)} headers", 1.0

    def _write_data(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self.worksheet: return False, "No active worksheet", 0.0
        data = params.get("data", task.data)
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, val in enumerate(row_data, 1):
                self.worksheet.cell(row=row_idx, column=col_idx, value=val)
        return True, f"Added {len(data)} rows of data", 1.0

    def _format_cells(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self.worksheet: return False, "No active worksheet", 0.0
        num_cols = len(task.columns)
        num_rows = len(task.data) + 1
        if params.get("bold", False):
            self.formatter.format_headers(self.worksheet, num_cols)
        if params.get("auto_width", False):
            self.formatter.auto_column_width(self.worksheet)
        self.formatter.apply_alternating_row_colors(self.worksheet, 2, num_rows, num_cols)
        return True, "Formatted cells", 1.0

    def _save_file(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self.workbook: return False, "No workbook to save", 0.0
        filename = params.get("filename", task.output_filename)
        self.current_file = os.path.join(self.output_dir, filename)
        self.workbook.save(self.current_file)
        return True, f"Saved to {self.current_file}", 1.0


class UIExecutor:
    """Execute plan intents via UI using PyAutoGUI with Screen Analyzer Action Guard."""

    def __init__(self, config: dict, logger: AgentLogger):
        self.config = config
        self.logger = logger
        self.screen_analyzer = ScreenAnalyzer(config, logger)
        self.action_delay = config.get("executor", {}).get("ui_action_delay", 0.5)
        self.load_wait = config.get("executor", {}).get("ui_load_wait", 5.0)
        self.save_wait = config.get("executor", {}).get("ui_save_wait", 2.0)
        pyautogui.PAUSE = self.action_delay

    def _action_guard(self, expected_file: Optional[str] = None) -> bool:
        """Action Guard Layer: Ensures UI is in safe state before typing/clicking."""
        if not self.screen_analyzer.enabled:
            return True

        if not self.screen_analyzer.is_excel_open_and_active():
            self.logger.warning("Action Guard Failed: Excel not active.")
            return False

        if expected_file and not self.screen_analyzer.is_correct_file_open(expected_file):
            self.logger.warning(f"Action Guard Failed: Expected file {expected_file} not open.")
            return False

        return True

    def execute_step(self, step: PlanStep, task: StructuredTask) -> Tuple[bool, str, float]:
        intent = step.intent
        params = step.params
        self.logger.info(f"[UI] Executing intent: {intent} → {step.target}")

        try:
            if intent == "create_workbook":
                return self._create_workbook(params)
            elif intent == "write_headers":
                return self._write_headers(params, task)
            elif intent == "write_data":
                return self._write_data(params, task)
            elif intent == "format_cells":
                return self._format_cells(params, task)
            elif intent == "save_file":
                return self._save_file(params, task)
            else:
                return False, f"Unknown intent: {intent}", 0.0
        except Exception as e:
            error_msg = f"UI execution error in '{intent}': {str(e)}"
            self.logger.error(error_msg)
            return False, error_msg, 0.0

    def _create_workbook(self, params: dict) -> Tuple[bool, str, float]:
        # Open Excel via OS command
        if platform.system() == "Windows":
            os.system("start excel")
        elif platform.system() == "Darwin":
            os.system("open -a 'Microsoft Excel'")
        else:
            self.logger.error("OS not supported for UI Excel automation")
            return False, "Unsupported OS", 0.0
            
        # Wait for Excel to load
        time.sleep(self.load_wait)
        self.screen_analyzer.wait_for_ui_change(timeout=self.load_wait)

        if not self._action_guard():
            return False, "Excel failed to open or activate properly", 0.0

        # Press Enter or Ctrl+N for new workbook
        pyautogui.hotkey("ctrl", "n")
        time.sleep(1)
        return True, "Opened new Excel UI workbook", 0.9

    def _write_headers(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self._action_guard(): return False, "Action Guard failed", 0.0
        
        # Assume A1 is selected after new workbook
        columns = params.get("columns", task.columns)
        for header in columns:
            pyautogui.write(str(header))
            pyautogui.press("tab")
        pyautogui.press("enter")
        return True, f"Typed {len(columns)} headers", 0.8

    def _write_data(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self._action_guard(): return False, "Action Guard failed", 0.0
        
        data = params.get("data", task.data)
        for row in data:
            for val in row:
                pyautogui.write(str(val))
                pyautogui.press("tab")
            pyautogui.press("enter")
        return True, f"Typed {len(data)} rows", 0.7

    def _format_cells(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self._action_guard(): return False, "Action Guard failed", 0.0
        # Simplistic UI formatting: Ctrl+A, then format as table (Ctrl+T)
        pyautogui.hotkey("ctrl", "a")
        time.sleep(0.5)
        pyautogui.hotkey("ctrl", "t")
        time.sleep(0.5)
        pyautogui.press("enter")
        return True, "Applied UI formatting", 0.8

    def _save_file(self, params: dict, task: StructuredTask) -> Tuple[bool, str, float]:
        if not self._action_guard(): return False, "Action Guard failed", 0.0
        
        filename = params.get("filename", task.output_filename)
        # F12 for Save As in Excel
        pyautogui.press("f12")
        time.sleep(1.0)
        self.screen_analyzer.wait_for_ui_change(timeout=2.0)
        
        abs_path = os.path.abspath(os.path.join(self.config.get("paths", {}).get("output_dir", "output"), filename))
        pyautogui.write(abs_path)
        time.sleep(0.5)
        pyautogui.press("enter")
        
        # Wait for save dialog to close
        time.sleep(self.save_wait)
        self.screen_analyzer.wait_for_ui_change(timeout=self.save_wait)
        
        # Confirm replace if it asks
        pyautogui.press("left")
        pyautogui.press("enter")
        
        return True, f"UI Saved to {abs_path}", 0.9


class ExecutorModule:
    """Orchestrates execution switching between Direct and UI modes."""

    def __init__(self, config: dict, logger: AgentLogger, event_bus=None):
        self.config = config
        self.logger = logger
        self.event_bus = event_bus
        self.direct_executor = DirectExecutor(config, logger)
        self.ui_executor = UIExecutor(config, logger)
        self.visual_executor = VisualExecutor(config, logger)
        self.execution_timeout = config.get("executor", {}).get("execution_timeout", 30)

    def _execute_with_timeout(self, executor, step: PlanStep, task: StructuredTask) -> Tuple[bool, str, float]:
        """Wrapper to enforce execution timeout on steps."""
        import concurrent.futures
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            future = pool.submit(executor.execute_step, step, task)
            try:
                # Wait for the result with a timeout
                return future.result(timeout=self.execution_timeout)
            except concurrent.futures.TimeoutError:
                self.logger.error(f"Execution Timeout! Step {step.step_id} exceeded {self.execution_timeout}s.")
                return False, f"Timeout after {self.execution_timeout} seconds.", 0.0
            except Exception as e:
                return False, str(e), 0.0

    def execute_plan(self, plan: Plan, task: StructuredTask, state=None) -> Tuple[bool, str]:
        self.logger.info(f"Executing plan ({plan.execution_mode.value}) with {len(plan.steps)} steps")
        
        overall_confidence = 1.0

        for idx, step in enumerate(plan.steps):
            if step.status == "done":
                # Idempotent check: if it's already done and it is idempotent, we can skip it.
                if step.is_idempotent:
                    self.logger.debug(f"Skipping idempotent step {step.step_id} which is already done.")
                    continue

            # Kill switch check
            if state and hasattr(state, "kill_switch"):
                from models.data_models import KillSwitchState
                if state.kill_switch == KillSwitchState.HARD_STOP:
                    return False, "Hard Stop Triggered by User. Aborting execution."
                if state.kill_switch == KillSwitchState.SOFT_STOP:
                    self.logger.info("Soft Stop requested. Finishing execution early.")
                    return False, "Soft Stop Triggered. execution halted gracefully."

            step.status = "running"
            
            # Observability Timeline Update
            if self.event_bus:
                self.event_bus.publish("timeline_update", {
                    "step": idx + 1, "total": len(plan.steps), "intent": step.intent, 
                    "status": "Running", "mode": plan.execution_mode.value
                })
            
            # Phase 1: Reliable Execution
            if plan.execution_mode == ExecutionMode.VISUAL:
                if step.intent == "save_file":
                    self.logger.info(f"Skipping save_file in Phase 1 for VISUAL mode")
                    continue
                executor = self.direct_executor
            else:
                executor = self.direct_executor if plan.execution_mode == ExecutionMode.DIRECT else self.ui_executor
                
            success, msg, conf = self._execute_with_timeout(executor, step, task)

            step.execution_confidence = conf
            overall_confidence = min(overall_confidence, conf)

            if self.event_bus:
                self.event_bus.publish("update_confidence", {"execution": overall_confidence})

            if success:
                step.status = "done"
                self.logger.step(idx + 1, len(plan.steps), f"Success: {msg}")
                if self.event_bus:
                    self.event_bus.publish("timeline_update", {
                        "step": idx + 1, "intent": step.intent, "status": "Success", "mode": plan.execution_mode.value
                    })
            else:
                step.status = "failed"
                step.error = msg
                self.logger.error(f"Step {idx + 1} failed: {msg}")
                if self.event_bus:
                    self.event_bus.publish("timeline_update", {
                        "step": idx + 1, "intent": step.intent, "status": "Failed", "error": msg, "mode": plan.execution_mode.value
                    })
                return False, msg

        # Phase 2: Visual Replay (Only if Phase 1 was fully successful)
        if plan.execution_mode == ExecutionMode.VISUAL:
            self.logger.info("Phase 1 Complete. Initiating Phase 2: Visual Replay")
            
            # Replay Data Immutability Guarantee
            replay_data = tuple(tuple(row) for row in task.data)
            
            try:
                self.visual_executor.replay(task, replay_data)
            except Exception as e:
                self.logger.warning(f"Visual replay failed, but data is safe. Error: {e}")

            # Phase 3: Silent Backend Save
            self.logger.info("Phase 3: Saving backend workbook...")
            save_step = next((s for s in plan.steps if s.intent == "save_file"), None)
            if save_step:
                success, msg, conf = self._execute_with_timeout(self.direct_executor, save_step, task)
                if success:
                    save_step.status = "done"
                else:
                    self.logger.error(f"Backend save failed: {msg}")

        return True, "All steps completed successfully"
