"""
Verifier Module — Validate output Excel file.
"""

import os
import openpyxl
from core.logger import AgentLogger
from models.data_models import VerificationResult, StructuredTask


class VerifierModule:
    """Verifies the correctness of the generated Excel file."""

    def __init__(self, config: dict, logger: AgentLogger):
        self.config = config
        self.logger = logger
        self.verification_config = config.get("verification", {})
        self.confidence_threshold = self.verification_config.get("confidence_threshold", 0.8)

    def verify(self, output_path: str, task: StructuredTask) -> VerificationResult:
        self.logger.info(f"Verifying output file: {output_path}")

        result = VerificationResult()
        
        if not os.path.exists(output_path):
            result.issues.append("File not found.")
            return result

        try:
            wb = openpyxl.load_workbook(output_path, data_only=True)
            ws = wb.active
        except Exception as e:
            result.issues.append(f"Cannot read Excel file: {e}")
            return result

        total_checks = 0
        passed_checks = 0

        # Check headers
        if self.verification_config.get("check_headers", True):
            total_checks += 1
            headers_match = self._verify_headers(ws, task.columns)
            if headers_match:
                passed_checks += 1
            else:
                result.issues.append("format issue: Headers do not match expected columns.")

        # Check row count
        if self.verification_config.get("check_row_count", True):
            total_checks += 1
            # -1 for header row
            actual_rows = max(0, ws.max_row - 1)
            expected_rows = len(task.data)
            
            if actual_rows >= expected_rows:
                passed_checks += 1
            else:
                result.issues.append(f"data mismatch: Expected {expected_rows} rows, found {actual_rows}.")

        # Calculate confidence
        if total_checks > 0:
            result.verification_confidence = passed_checks / total_checks
        else:
            result.verification_confidence = 1.0

        result.passed = len(result.issues) == 0 and result.verification_confidence >= self.confidence_threshold
        
        self.logger.info(
            f"Verification {'PASSED' if result.passed else 'FAILED'}. "
            f"Confidence: {result.verification_confidence:.2f}, Issues: {len(result.issues)}"
        )
        return result

    def _verify_headers(self, ws, expected_columns: list) -> bool:
        if not expected_columns:
            return True
            
        actual_headers = []
        for col in range(1, len(expected_columns) + 1):
            val = ws.cell(row=1, column=col).value
            actual_headers.append(str(val) if val is not None else "")
            
        # Case insensitive match
        actual_lower = [h.lower().strip() for h in actual_headers]
        expected_lower = [h.lower().strip() for h in expected_columns]
        
        return actual_lower == expected_lower
