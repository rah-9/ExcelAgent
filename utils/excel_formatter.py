"""
Excel Formatter — openpyxl-based formatting utilities.
"""

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Optional


class ExcelFormatter:
    """Apply professional formatting to openpyxl worksheets."""

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, config: dict = None):
        self.config = config or {}
        excel_config = self.config.get("executor", {}).get("excel", {})

        self.header_fill = PatternFill(
            start_color=excel_config.get("header_fill_color", "4472C4"),
            end_color=excel_config.get("header_fill_color", "4472C4"),
            fill_type="solid",
        )
        self.header_font = Font(
            name=excel_config.get("header_font", "Calibri"),
            bold=excel_config.get("header_bold", True),
            color=excel_config.get("header_font_color", "FFFFFF"),
            size=excel_config.get("header_font_size", 12),
        )
        self.data_font = Font(
            name=excel_config.get("default_font", "Calibri"),
            size=excel_config.get("data_font_size", 11),
        )
        self.header_alignment = Alignment(
            horizontal=excel_config.get("alignment_horizontal", "center"),
            vertical="center",
            wrap_text=True,
        )
        self.data_alignment = Alignment(
            horizontal=excel_config.get("alignment_horizontal", "center"),
            vertical="center",
        )

    def format_headers(self, ws, num_columns: int, row: int = 1):
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.THIN_BORDER

    def format_data_cells(self, ws, start_row: int, end_row: int, num_columns: int):
        for row in range(start_row, end_row + 1):
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.THIN_BORDER

    def auto_column_width(self, ws, min_width: int = 10, max_width: int = 50):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    def apply_alternating_row_colors(self, ws, start_row: int, end_row: int,
                                   num_columns: int, color: str = "D9E2F3"):
        alt_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in range(start_row, end_row + 1):
            if (row - start_row) % 2 == 1:
                for col in range(1, num_columns + 1):
                    ws.cell(row=row, column=col).fill = alt_fill

    def apply_full_formatting(self, ws, num_columns: int, num_rows: int):
        self.format_headers(ws, num_columns)
        if num_rows > 1:
            self.format_data_cells(ws, 2, num_rows, num_columns)
            self.apply_alternating_row_colors(ws, 2, num_rows, num_columns)
        self.auto_column_width(ws)
